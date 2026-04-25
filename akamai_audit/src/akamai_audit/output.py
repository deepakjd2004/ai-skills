from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook


def write_outputs(result: dict[str, Any], output_dir: Path, base_name: str = "sample_output") -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    with (output_dir / "result.json").open("w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)

    for table_name, rows in result.items():
        if not isinstance(rows, list) or not rows:
            continue
        if not isinstance(rows[0], dict):
            continue

        file_path = output_dir / f"{table_name}.csv"
        with file_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    safe_base_name = _sanitize_filename(base_name) or "sample_output"
    _write_sample_workbook(result, output_dir / f"{safe_base_name}.xlsx")


def _write_sample_workbook(result: dict[str, Any], file_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _add_table_sheet(
        wb,
        "Contracts",
        ["Contract ID", "Product Id", "Product Name"],
        _map_rows(result.get("contracts", []), ["contract_id", "product_id", "product_name"]),
    )
    _add_table_sheet(
        wb,
        "Groups",
        ["Contract ID", "Group ID", "Group Name", "Parent Group ID", "Parent Group Name"],
        _map_rows(
            result.get("groups", []),
            ["contract_id", "group_id", "group_name", "parent_group_id", "parent_group_name"],
        ),
    )
    _add_table_sheet(
        wb,
        "CP Codes",
        ["Contract ID", "CP Code", "CP Code Name", "CP Code Products"],
        _map_rows(
            result.get("cpcodes", []),
            ["contract_id", "cpcode_id", "cpcode_name", "cpcode_products"],
        ),
    )
    _add_table_sheet(
        wb,
        "Cloudlets",
        ["Policy ID", "Policy Name", "Policy Type", "Cloudlet Type", "Associated Properties", "# of Properties"],
        _map_rows(
            result.get("cloudlets", []),
            [
                "policy_id",
                "policy_name",
                "policy_type",
                "cloudlet_type",
                "associated_properties",
                "property_count",
            ],
        ),
    )

    property_rows = []
    for row in result.get("properties", []):
        property_rows.append(
            [
                row.get("property_id", ""),
                row.get("property_name", ""),
                row.get("group_id", ""),
                row.get("contract_id", ""),
                row.get("production_status", ""),
                row.get("staging_status", ""),
                row.get("property_version", ""),
                row.get("updated_date", ""),
                row.get("note", ""),
                row.get("updated_by_user", ""),
                ",".join(row.get("cpcodes", [])) if isinstance(row.get("cpcodes"), list) else row.get("cpcodes", ""),
                ",".join(row.get("origin_hostnames", [])) if isinstance(row.get("origin_hostnames"), list) else row.get("origin_hostnames", ""),
            ]
        )
    _add_table_sheet(
        wb,
        "Property",
        [
            "Property ID",
            "Property Name",
            "Group",
            "Contract",
            "Prod Status",
            "Staging Status",
            "Version",
            "Last Modified",
            "Notes",
            " User",
            "CP Codes",
            "Origin Servers",
        ],
        property_rows,
    )

    _add_table_sheet(
        wb,
        "Property Behaviors",
        ["Configuration | Version | Format | Product", "Behavior", "In Use?", "Count"],
        _map_rows(
            result.get("property_behaviors", []),
            ["Configuration | Version | Format | Product", "Behavior", "In Use?", "Count"],
        ),
    )
    _add_table_sheet(
        wb,
        "Property Hostnames",
        ["Hostnames DNS Resolution", "Origin Hostnames DNS Resolution"],
        _map_rows(
            result.get("property_hostnames", []),
            ["Hostnames DNS Resolution", "Origin Hostnames DNS Resolution"],
        ),
    )

    _add_traffic_summary_sheet(wb, result)
    _add_url_traffic_hits_sheet(wb, result)
    _add_table_sheet(
        wb,
        "Response Codes",
        ["Response Code", "Edge Hits", "Edge Hits %", "Origin Hits", "Origin Hits %"],
        _map_rows(
            result.get("response_codes", []),
            ["response_code", "edge_hits", "edge_hits_percent", "origin_hits", "origin_hits_percent"],
        ),
    )
    _add_table_sheet(
        wb,
        "302 URLs",
        ["URL", "302 Edge Hits"],
        _map_rows(result.get("url_302", []), ["url", "hits"]),
    )
    _add_table_sheet(
        wb,
        "304 URLs",
        ["URL", "304 Edge Hits"],
        _map_rows(result.get("url_304", []), ["url", "hits"]),
    )
    _add_table_sheet(
        wb,
        "404 URLs",
        ["URL", "404 Edge Hits"],
        _map_rows(result.get("url_404", []), ["url", "hits"]),
    )

    wb.save(file_path)


def _add_table_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _add_traffic_summary_sheet(wb: Workbook, result: dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Traffic Summary")
    ws.append(
        [
            "CP Code",
            "CP Code Name",
            "Offload",
            "Edge Hits",
            "Origin Hits",
            None,
            "CP Code",
            "CP Code Name",
            "Offload",
            "Edge Bytes",
            "Origin Bytes",
        ]
    )

    hits_rows = _map_rows(
        result.get("traffic_summary_hits", []),
        ["cpcode_id", "cpcode_name", "offload_percent", "edge_hits", "origin_hits"],
    )
    bytes_rows = _map_rows(
        result.get("traffic_summary_bytes", []),
        ["cpcode_id", "cpcode_name", "offload_percent", "edge_bytes", "origin_bytes"],
    )

    total = max(len(hits_rows), len(bytes_rows))
    for i in range(total):
        left = hits_rows[i] if i < len(hits_rows) else ["", "", "", "", ""]
        right = bytes_rows[i] if i < len(bytes_rows) else ["", "", "", "", ""]
        ws.append(left + [None] + right)


def _add_url_traffic_hits_sheet(wb: Workbook, result: dict[str, Any]) -> None:
    ws = wb.create_sheet(title="URL Traffic Hits")
    ws.append(["URL", "Edge Hits", "Offload", "Origin Hits", "File Extension", None, None, "", None, None, None, None])

    rows = _map_rows(
        result.get("url_traffic_hits", []),
        ["url", "edge_hits", "offload_percent", "origin_hits", "file_extension"],
    )
    for row in rows:
        ws.append(row)

    ws.cell(row=1, column=8).value = "File Extension"
    ws.cell(row=1, column=9).value = "SUM of Edge Hits"
    ws.cell(row=1, column=10).value = "SUM of Edge Hits %"
    ws.cell(row=1, column=11).value = "SUM of Origin Hits"
    ws.cell(row=1, column=12).value = "Offload"

    summary_rows = _map_rows(
        result.get("url_traffic_extension_summary", []),
        ["file_extension", "edge_hits_sum", "edge_hits_percent", "origin_hits_sum", "offload_percent"],
    )
    row_idx = 2
    for row in summary_rows:
        for col_idx, value in enumerate(row, start=8):
            ws.cell(row=row_idx, column=col_idx).value = value
        row_idx += 1


def _map_rows(rows: list[dict[str, Any]], keys: list[str]) -> list[list[Any]]:
    mapped: list[list[Any]] = []
    for row in rows or []:
        mapped.append([row.get(key, "") for key in keys])
    return mapped


def _sanitize_filename(value: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_")
    return "".join(ch if ch in allowed else "_" for ch in value).strip("_")
